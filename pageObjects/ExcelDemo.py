import openpyxl


class exceldemo:
    test_exceldemo_data = [{"FirstName": "Shruti", "Gender": "Female", "LastName": "Soni"},
                           {"FirstName": "Divyang", "Gender": "Male", "LastName": "Soni"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("D:\\pytest\\ExcelDemo.xlsx")
        sheet = book.active
        Dict = {}
        cell = sheet.cell(row=1, column=2)

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
